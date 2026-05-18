"""Migration one-shot : xlsx existant → SQLite.

Lit data/source/candidatures_alternance_AI_Engineer.xlsx, qui contient 6 tables :
  - Table_1 : les vraies offres scrapées (~194 lignes) → table `offers`
  - Tables 2-6 : entreprises "cibles candidature spontanée" → JSON pour phase 2

Sépare les vrais statuts des notes "Comment postuler", nettoie le mojibake encoding.

Le xlsx n'est PAS modifié (read-only).

⚠️ SCRIPT DESTRUCTIF : fait `DELETE FROM offers` avant l'import. Pour éviter
qu'un user perde 1000+ offres scrapées en relançant le quickstart par réflexe,
le garde-fou `_check_safe_to_wipe` refuse de tourner si la table `offers`
est déjà peuplée — il faut passer `--force` explicitement.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from backend.db import db, init_schema, make_dedup_key
from backend.models import VALID_STATUSES, label_for_score

ROOT = Path(__file__).resolve().parent.parent
COMPANIES_JSON = ROOT / "data" / "companies_spontaneous_extracted.json"

XLSX_PATH = ROOT / "data" / "source" / "candidatures_alternance_AI_Engineer.xlsx"

# Colonnes du xlsx (1-indexed, basé sur l'inspection)
COL_MATCH = 2
COL_TITRE = 3
COL_ENTREPRISE = 4
COL_VILLE = 5
COL_DEPARTEMENT = 6
COL_SOURCE = 7
COL_URL = 8                # cellule contient "Voir l'offre" + hyperlien
COL_DATE_PUB = 9
COL_REMOTE = 10
COL_DATE_CAND = 11
COL_STATUT = 12
COL_DATE_RELANCE = 13
COL_DATE_ENTRETIEN = 14


# Mapping mojibake → caractère propre (encoding cp1252→utf8 raté)
MOJIBAKE_MAP = {
    "é": "é", "è": "è", "ê": "ê", "à": "à", "ç": "ç",
    "ô": "ô", "î": "î", "ï": "ï", "û": "û", "ù": "ù",
    "É": "É", "À": "À", "Ç": "Ç", "Ô": "Ô", "Ê": "Ê",
    "œ": "œ", "Œ": "Œ", "€": "€", "’": "'", "”": "\"", "“": "\"",
    "…": "...", "–": "-", "—": "-",
    # cas garbled (caractère de remplacement)
    "�": "",
}

# Normalisations pour les statuts (clean → canonique)
STATUT_CANONIQUE = {
    "Postulé": "Postulé",
    "Postule": "Postulé",
    "POSTULÉ": "Postulé",
    "Refusé": "Refusé",
    "Refuse": "Refusé",
    "REFUSÉ": "Refusé",
    "Relancé": "Relancé",
    "Relance": "Relancé",
    "Entretien": "Entretien",
    "Test technique": "Test technique",
    "Accepté": "Accepté",
    "Accepte": "Accepté",
    "Sans réponse": "Sans réponse",
    "Sans reponse": "Sans réponse",
    "Abandonné": "Abandonné",
    "Abandonne": "Abandonné",
    "A postuler": None,         # vide = pas encore postulé
    "À postuler": None,
}


def fix_mojibake(text: str | None) -> str | None:
    if not text:
        return text
    s = str(text)
    for bad, good in MOJIBAKE_MAP.items():
        s = s.replace(bad, good)
    return s.strip() or None


def parse_statut_cell(raw: str | None) -> tuple[str | None, str | None]:
    """Sépare la colonne Statut polluée en (status, application_method).

    Retourne (statut_propre, methode_candidature). L'un OU l'autre OU les deux
    peuvent être None.
    """
    if not raw:
        return (None, None)
    cleaned = fix_mojibake(raw)
    if cleaned is None:
        return (None, None)

    # Match exact (avec ou sans variations d'accents/casse)
    if cleaned in STATUT_CANONIQUE:
        return (STATUT_CANONIQUE[cleaned], None)

    # Sinon c'est une note "Comment postuler"
    return (None, cleaned)


def cell_url(cell) -> str | None:
    """Récupère l'URL réelle d'une cellule (hyperlink override si présent)."""
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    v = cell.value
    if v and isinstance(v, str) and v.startswith(("http://", "https://")):
        return v
    return None


def cell_str(cell) -> str | None:
    v = cell.value
    if v is None:
        return None
    s = str(v).strip()
    return fix_mojibake(s) or None


def cell_int(cell) -> int | None:
    v = cell.value
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def cell_date(cell) -> str | None:
    """Retourne une date ISO ou None."""
    v = cell.value
    if v is None or v == "":
        return None
    # Si c'est déjà un objet date/datetime
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    s = str(v).strip()
    if not s:
        return None
    # déjà en ISO ?
    import re
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    # dd/mm/yyyy
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    return s  # leave as-is


def extract_companies_tables(ws, tables: dict) -> list[dict]:
    """Tables 2-6 : entreprises 'candidature spontanée'. Sauvées brutes."""
    result = []
    for name in ("Table_2", "Table_3", "Table_4", "Table_5", "Table_6"):
        if name not in tables:
            continue
        ref = tables[name]
        col_min, row_min, col_max, row_max = range_boundaries(ref)
        # 1re ligne = headers
        headers = [cell_str(ws.cell(row_min, c)) for c in range(col_min, col_max + 1)]
        for r in range(row_min + 1, row_max + 1):
            values = {}
            non_empty = False
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = cell_str(ws.cell(r, col_min + i))
                values[h] = v
                if v:
                    non_empty = True
            if non_empty:
                values["_source_table"] = name
                result.append(values)
    return result


def _count_existing_offers() -> int:
    """Compte les offres en DB. Init le schéma si nécessaire (idempotent)."""
    init_schema()
    with db() as conn:
        return int(conn.execute("SELECT COUNT(*) FROM offers").fetchone()[0])


def _check_safe_to_wipe(*, force: bool) -> tuple[bool, str | None]:
    """Garde-fou destructif.

    Refuse de wiper la table `offers` si elle est déjà peuplée, sauf si
    `force=True`. Retourne `(safe, error_message)` :
    - `(True, None)` : OK pour wiper (DB neuve OU --force).
    - `(False, msg)` : refus, message d'erreur à afficher.

    Pure-function, testable sans toucher au xlsx.
    """
    existing = _count_existing_offers()
    if existing == 0:
        return True, None
    if force:
        return True, None
    # SAFE (B608) : message d'erreur user-facing (printé sur stderr), pas
    # une query SQL. Bandit fait du keyword matching sur "DELETE FROM" dans
    # les littéraux sans distinguer le contexte. `existing` est un `int`
    # retourné par `_count_existing_offers()`, pas un input user.
    msg = (
        "REFUS : la table 'offers' contient déjà {n} ligne(s).\n"
        "Ce script fait \"DELETE FROM offers\" avant l'import xlsx — perte\n"  # nosec B608
        "de toutes les offres scrapées si tu continues.\n"
        "\n"
        "Si tu es SÛR de vouloir réinitialiser la DB depuis le xlsx :\n"
        "    python -m backend.migrate_xlsx --force\n"
        "\n"
        "Sinon, tu cherches probablement :\n"
        "    python cli.py scrape all      # rafraîchir les offres\n"
        "    python -m backend             # démarrer l'app sans rien casser"
    ).format(n=existing)
    return False, msg


def run(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m backend.migrate_xlsx",
        description=(
            "Import one-shot xlsx historique → SQLite. DESTRUCTIF : fait "
            "DELETE FROM offers avant l'import. Requiert --force si la "
            "table est déjà peuplée."
        ),
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help=(
            "Confirme explicitement que tu veux ÉCRASER la table offers "
            "actuelle (perte des données scrapées). Requis dès que la "
            "table contient au moins 1 ligne."
        ),
    )
    args = parser.parse_args(argv)

    # ORDRE CRITIQUE : la garde anti-wipe AVANT le check xlsx. Si l'utilisateur
    # a 1000 offres scrapées + xlsx introuvable, on veut afficher "REFUS, utilise
    # --force" plutôt que "xlsx introuvable" — ça protège ses données même
    # avant d'évaluer le scénario d'import.
    safe, refusal = _check_safe_to_wipe(force=args.force)
    if not safe:
        print(refusal, file=sys.stderr)
        return 2

    if not XLSX_PATH.exists():
        print(f"ERREUR : xlsx introuvable : {XLSX_PATH}", file=sys.stderr)
        return 1

    print(f"Lecture de {XLSX_PATH}...")
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Candidatures"]

    # openpyxl 3.x : ws.tables est dict[name -> ref_string]
    tables = dict(ws.tables.items())
    if "Table_1" not in tables:
        print("ERREUR : Table_1 introuvable", file=sys.stderr)
        return 1
    _, t1_min_row, _, t1_max_row = range_boundaries(tables["Table_1"])
    print(f"Table_1 (offres) : lignes {t1_min_row} à {t1_max_row}")

    # Init du schéma (déjà fait par _check_safe_to_wipe → no-op ici)
    init_schema()

    # Lecture des lignes
    rows = []
    statuts_split = {"clean": 0, "method": 0, "vide": 0}
    for r in range(t1_min_row + 1, t1_max_row + 1):
        titre = cell_str(ws.cell(r, COL_TITRE))
        if not titre:
            continue

        entreprise = cell_str(ws.cell(r, COL_ENTREPRISE))
        ville = cell_str(ws.cell(r, COL_VILLE))
        url = cell_url(ws.cell(r, COL_URL))
        statut_raw = cell_str(ws.cell(r, COL_STATUT))
        statut_clean, methode = parse_statut_cell(statut_raw)

        if statut_clean:
            statuts_split["clean"] += 1
        elif methode:
            statuts_split["method"] += 1
        else:
            statuts_split["vide"] += 1

        match_score = cell_int(ws.cell(r, COL_MATCH))
        match_label = label_for_score(match_score)

        rows.append({
            "title": titre,
            "company": entreprise,
            "city": ville,
            "department": cell_str(ws.cell(r, COL_DEPARTEMENT)),
            "source": cell_str(ws.cell(r, COL_SOURCE)),
            "url": url,
            "date_published": cell_date(ws.cell(r, COL_DATE_PUB)),
            "remote": cell_str(ws.cell(r, COL_REMOTE)),
            "date_applied": cell_date(ws.cell(r, COL_DATE_CAND)),
            "status": statut_clean,
            "application_method": methode,
            "date_followup": cell_date(ws.cell(r, COL_DATE_RELANCE)),
            "date_interview": cell_date(ws.cell(r, COL_DATE_ENTRETIEN)),
            "match_score": match_score,
            "match_label": match_label,
            "contract_type": "Alternance",
            # IMPORTANT : la ville fait partie de la clé (audit bug : Paris vs
            # Toulouse = 2 offres distinctes). Sans ce 3e arg, l'import xlsx
            # générait des `"titre|entreprise|"` qui ne matchaient PAS les
            # `"titre|entreprise|paris"` produits par les scrapers — doublons
            # silencieux au prochain scrape Hellowork/WTTJ.
            "dedup_key": make_dedup_key(titre, entreprise, ville),
        })

    print(f"Lignes lues : {len(rows)}")
    print(f"  Statut clean : {statuts_split['clean']}")
    print(f"  Notes 'Comment postuler' : {statuts_split['method']}")
    print(f"  Vides : {statuts_split['vide']}")

    # Insertion (truncate puis insert pour idempotence de la migration)
    with db() as conn:
        # On vide d'abord pour pouvoir re-run la migration (et reset AUTOINCREMENT)
        conn.execute("DELETE FROM offers")
        conn.execute("DELETE FROM sqlite_sequence WHERE name = 'offers'")

        sql = """
            INSERT INTO offers (
                title, company, city, department, source, url,
                date_published, remote, contract_type,
                match_score, match_label,
                status, application_method,
                date_applied, date_followup, date_interview,
                dedup_key, scraped_at
            ) VALUES (
                :title, :company, :city, :department, :source, :url,
                :date_published, :remote, :contract_type,
                :match_score, :match_label,
                :status, :application_method,
                :date_applied, :date_followup, :date_interview,
                :dedup_key, CURRENT_TIMESTAMP
            )
        """
        inserted = 0
        skipped_url_dup = 0
        seen_urls: set[str] = set()
        seen_keys: set[str] = set()
        for row in rows:
            # Dédoublonnage intra-xlsx (par sécurité)
            if row["url"] and row["url"] in seen_urls:
                skipped_url_dup += 1
                continue
            if row["dedup_key"] in seen_keys and not row["url"]:
                skipped_url_dup += 1
                continue
            if row["url"]:
                seen_urls.add(row["url"])
            seen_keys.add(row["dedup_key"])
            conn.execute(sql, row)
            inserted += 1

    print(f"Insérées : {inserted}")
    print(f"Doublons skippés : {skipped_url_dup}")

    # Stats finales
    with db() as conn:
        total = conn.execute("SELECT COUNT(*) FROM offers").fetchone()[0]
        with_score = conn.execute("SELECT COUNT(*) FROM offers WHERE match_score IS NOT NULL").fetchone()[0]
        by_status = conn.execute("SELECT status, COUNT(*) FROM offers GROUP BY status ORDER BY 2 DESC").fetchall()
        by_label = conn.execute("SELECT match_label, COUNT(*) FROM offers GROUP BY match_label ORDER BY 2 DESC").fetchall()

    print(f"\nDB finale : {total} offres, {with_score} avec score")
    print("Statuts :")
    for s, c in by_status:
        print(f"  {s or '(aucun)'} : {c}")
    print("Match label :")
    for l, c in by_label:
        print(f"  {l or '(aucun)'} : {c}")

    # --- Tables 2-6 : extraction des entreprises 'candidature spontanée' vers JSON ---
    companies = extract_companies_tables(ws, tables)
    COMPANIES_JSON.parent.mkdir(parents=True, exist_ok=True)
    COMPANIES_JSON.write_text(
        json.dumps(companies, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\nEntreprises 'candidature spontanee' extraites : {len(companies)} -> {COMPANIES_JSON}")

    return 0


if __name__ == "__main__":
    sys.exit(run())
